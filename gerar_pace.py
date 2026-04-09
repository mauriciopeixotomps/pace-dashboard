import openpyxl, json, sys, re as re_mod
from datetime import datetime, date
from collections import defaultdict
import os as _os

_script_dir = _os.path.dirname(_os.path.abspath(__file__))
UPLOAD_DIR  = _os.path.join(_script_dir, '..', 'uploads')
OUTPUT_HTML = _os.path.join(_script_dir, 'pace_abril_2026.html')

META_RE  = 600   # CC1 Realizadas
META_AG  = 903   # CC1 Agendadas

WORKING_DAYS = [
    date(2026,4,1), date(2026,4,2), date(2026,4,3),
    date(2026,4,6), date(2026,4,7), date(2026,4,8),
    date(2026,4,9), date(2026,4,10),
    date(2026,4,13), date(2026,4,14), date(2026,4,15), date(2026,4,16), date(2026,4,17),
    date(2026,4,22), date(2026,4,23), date(2026,4,24),
    date(2026,4,27), date(2026,4,28), date(2026,4,29), date(2026,4,30)
]
def get_today_wd():
    t = date.today()
    past = [d for d in WORKING_DAYS if d <= t]
    return past[-1] if past else WORKING_DAYS[0]
TODAY = get_today_wd()

# LEADS_PLANILHA é construído após os helpers (veja abaixo)

# ── Helpers ───────────────────────────────────────────────────────────────
def get_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    s = str(val).strip()
    try: return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except: return None

def clean(v): return str(v).strip() if v and str(v) not in ('None','nan','') else ''

# Produto GS Partner para classificação de produto
GS_PARTNER_TEAM = {'samuel almeida', 'haynnã'}

# Closers que aparecem como Proprietário em deals mas NÃO são SDRs
CLOSER_ONLY_TEAM = {'samuel almeida', 'haynnã'}

def get_product(equipe, funil, etiqueta='', responsavel=''):
    # Verifica se o responsável é do time GS Partner
    if str(responsavel).lower().strip() in GS_PARTNER_TEAM: return 'GS'
    eq = str(equipe).upper() if equipe else ''
    fu = str(funil).upper() if funil else ''
    et = str(etiqueta).upper() if etiqueta else ''
    if 'STUDIO FISCAL' in eq or 'STUDIO FISCAL' in et: return 'SF'
    if 'STUDIO AGRO' in eq or ('AGRO' in et and 'FISCAL' not in et): return 'SA'
    if 'PJ360' in eq or 'PJ360' in et or 'TECNOLOGIA' in fu or 'EXPANS' in fu: return 'PJ360'
    if 'PARCERIAS' in eq or 'PARTNER' in fu or 'GS' in fu: return 'GS'
    if 'FISCAL' in fu: return 'SF'
    if 'FRANQUIA' in fu: return 'SA'
    return 'Outro'

def find_file(*patterns):
    """Tenta cada padrão em ordem; retorna o arquivo mais recente que casar."""
    for p in patterns:
        matches = [f for f in _os.listdir(UPLOAD_DIR) if p in f and f.endswith('.xlsx')]
        if matches:
            return _os.path.join(UPLOAD_DIR, sorted(matches)[-1])
    return None

def load_ws(path):
    if not path or not _os.path.exists(path): return None
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb.active

# ── Leads automático (1735: Negócio criado em) ────────────────────────────
def build_leads_planilha():
    ws_leads = load_ws(find_file('1735'))
    if not ws_leads:
        return {'2026-04-01': 139, '2026-04-02': 150, '2026-04-03': 160,
                '2026-04-06': 182, '2026-04-07': 88}
    raw = defaultdict(int)
    for row in ws_leads.iter_rows(min_row=2, values_only=True):
        d = get_date(row[5])  # Negócio - Negócio criado em
        if d and d.month == 4: raw[d] += 1
    wd_set = set(WORKING_DAYS)
    result = defaultdict(int)
    for d, cnt in raw.items():
        if d in wd_set:
            result[str(d)] += cnt
        else:
            nxt = next((w for w in WORKING_DAYS if w > d), None)
            if nxt: result[str(nxt)] += cnt
    return dict(result)

LEADS_PLANILHA = build_leads_planilha()

# ── Contadores ────────────────────────────────────────────────────────────
quals_by_day    = defaultdict(int)

# AG realizados (1718): SDR = col3 (Usuário responsável); Closer = col8 (Proprietário)
ag_by_day       = defaultdict(int)
ag_by_sdr       = defaultdict(int)
ag_by_closer    = defaultdict(int)
ag_by_product   = defaultdict(int)
ag_closer_prod  = defaultdict(lambda: defaultdict(int))
ag_sdr_prod     = defaultdict(lambda: defaultdict(int))

# RE realizadas (1719): Closer = col2 (Usuário responsável); SDR = col5 (Criador)
re_by_day       = defaultdict(int)
re_by_closer    = defaultdict(int)
re_by_sdr       = defaultdict(int)
re_by_product   = defaultdict(int)

# NS (1723): Closer = col4 (Usuário responsável); SDR = col5 (Criador)
ns_by_day       = defaultdict(int)
ns_by_closer    = defaultdict(int)
ns_by_sdr       = defaultdict(int)

# AG futuros (1722): SDR = col5 (Usuário responsável); Closer = col10 (Proprietário)
ag_fut_by_sdr    = defaultdict(int)
ag_fut_by_closer = defaultdict(int)
ag_fut_by_product= defaultdict(int)
ag_fut_by_day    = defaultdict(int)

# Perdidos (1721)
perdidos_by_day      = defaultdict(int)
perdidos_by_motivo   = defaultdict(int)
perdidos_by_campaign = defaultdict(int)
perdidos_by_etiqueta = defaultdict(int)
perdidos_daily_product = defaultdict(lambda: defaultdict(int))  # date → product → count

# ── Extrair dados ─────────────────────────────────────────────────────────

# Qualificações Realizadas (1736): col6=DataVenc, col3=SDR
ws = load_ws(find_file('1736', '1717'))
if ws:
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = get_date(row[6]) or get_date(row[5])
        if d and d.month == 4: quals_by_day[str(d)] += 1

# Agendamentos (1737, was 1718)
# col0=Etiqueta, col1=Equipe, col2=Funil, col3=SDR(Usuário resp.), col8=Closer(Proprietário), col12=DataVenc
ws = load_ws(find_file('1737', '1718'))
if ws:
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = get_date(row[12]) or get_date(row[11])
        if not d or d.month != 4: continue
        p     = get_product(row[1], row[2], row[0])
        sdr   = clean(row[3]) or 'N/A'
        closer= clean(row[8]) or clean(row[3]) or 'N/A'
        ag_by_day[str(d)] += 1
        ag_by_sdr[sdr] += 1
        ag_by_closer[closer] += 1
        ag_by_product[p] += 1
        ag_sdr_prod[sdr][p] += 1
        ag_closer_prod[closer][p] += 1

# Realizadas (1740, was 1719)
# col0=Equipe, col1=Funil, col2=Closer(Usuário resp.), col5=SDR(Criador), col8=DataVenc
ws = load_ws(find_file('1740', '1719'))
if ws:
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = get_date(row[8]) or get_date(row[10])
        if not d or d.month != 4: continue
        p      = get_product(row[0], row[1])
        closer = clean(row[2]) or 'N/A'
        sdr    = clean(row[5]) or 'N/A'
        re_by_day[str(d)] += 1
        re_by_closer[closer] += 1
        re_by_sdr[sdr] += 1
        re_by_product[p] += 1

# No Shows (1742, was 1723)
# col0=Equipe, col1=Funil, col4=Closer(Usuário resp.), col5=SDR(Criador), col8=DataVenc
ws = load_ws(find_file('1742', '1723'))
if ws:
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = get_date(row[8]) or get_date(row[7])
        if not d or d.month != 4: continue
        closer = clean(row[4]) or 'N/A'
        sdr    = clean(row[5]) or 'N/A'
        ns_by_day[str(d)] += 1
        ns_by_closer[closer] += 1
        ns_by_sdr[sdr] += 1

# AG Futuros (1741, was 1722)
# col0=Funil, col5=Closer(Usuário resp.=quem conduzirá), col10=SDR(Proprietário=dono do deal), col9=DataVenc
ws = load_ws(find_file('1741', '1722'))
if ws:
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = get_date(row[9]) or get_date(row[7])
        if not d or d.month != 4: continue
        p      = get_product('', row[0])
        closer = clean(row[5]) or 'N/A'
        sdr    = clean(row[10]) or 'N/A'
        # Não contar closers como SDR
        if sdr.lower() not in CLOSER_ONLY_TEAM:
            ag_fut_by_sdr[sdr] += 1
        ag_fut_by_closer[closer] += 1
        ag_fut_by_product[p] += 1
        ag_fut_by_day[str(d)] += 1

# Perdidos (1738, was 1721)
# col0=Etiqueta, col6=Motivo, col10=UTM campaign, col15=DataPerda
ws = load_ws(find_file('1738', '1721'))
if ws:
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = get_date(row[15])
        if not d or d.month != 4: continue
        motivo = clean(row[6]) or 'Não informado'
        etiq   = clean(row[0]) or 'Sem etiqueta'
        camp   = clean(row[10])
        p      = ('SF' if 'FISCAL' in etiq.upper() else
                  'SA' if 'AGRO' in etiq.upper() else
                  'PJ360' if 'PJ360' in etiq.upper() else 'Outro')
        perdidos_by_day[str(d)] += 1
        perdidos_by_motivo[motivo] += 1
        perdidos_by_etiqueta[etiq] += 1
        perdidos_daily_product[str(d)][p] += 1
        if camp: perdidos_by_campaign[camp] += 1

# ── Calcular metas derivadas ──────────────────────────────────────────────
total_quals      = sum(quals_by_day.values())
total_ag         = sum(ag_by_day.values())
total_re         = sum(re_by_day.values())
total_ns         = sum(ns_by_day.values())
total_leads      = sum(LEADS_PLANILHA.values())
total_perdidos   = sum(perdidos_by_day.values())
total_ag_fut     = sum(ag_fut_by_day.values())

qual_rate  = total_quals / total_leads if total_leads > 0 else 0.20
q2a_rate   = total_ag / total_quals if total_quals > 0 else 0.72
ag2re_rate = total_re / (total_ag + total_ns) if (total_ag + total_ns) > 0 else 0.66

META_LEADS = round(META_RE / (qual_rate * q2a_rate * ag2re_rate))

days_done  = len([d for d in WORKING_DAYS if d <= TODAY])
days_left  = len([d for d in WORKING_DAYS if d > TODAY])
total_days = len(WORKING_DAYS)

leads_pace_daily = total_leads / len(LEADS_PLANILHA) if LEADS_PLANILHA else 0
ag_pace_daily    = total_ag / days_done
re_pace_daily    = total_re / days_done
perdidos_pace    = total_perdidos / days_done

# AG Futuros: pipeline ideal = (RE + NS por dia) × 5 dias de cobertura
consumption_rate = (total_re + total_ns) / days_done if days_done > 0 else 0
ag_fut_ideal     = round(consumption_rate * 5)   # cobertura de 5 dias úteis

# Meta por closer: 70 RE realizadas no mês
META_CLOSER       = 70
closer_ideal_hoje = round(META_CLOSER * days_done / total_days)
# Projeção e ritmo por closer (combina RE + AG futuras)
closer_proj  = {}
closer_needed= {}
all_closers  = set(re_by_closer.keys()) | set(ag_fut_by_closer.keys())
for c in all_closers:
    re_done = re_by_closer.get(c, 0)
    fut     = ag_fut_by_closer.get(c, 0)
    closer_proj[c]   = re_done + fut
    needed = round((META_CLOSER - re_done) / days_left) if days_left > 0 else 0
    closer_needed[c] = max(0, needed)

leads_proj    = round(leads_pace_daily * total_days)
ag_proj       = round(ag_pace_daily * total_days)
re_proj       = round(re_pace_daily * total_days)
perdidos_proj = round(perdidos_pace * total_days)

leads_daily_needed = round((META_LEADS - total_leads) / days_left) if days_left > 0 else 0
ag_daily_needed    = round((META_AG - total_ag) / days_left) if days_left > 0 else 0
re_daily_needed    = round((META_RE - total_re) / days_left) if days_left > 0 else 0

leads_ideal  = round(META_LEADS * days_done / total_days)
ag_ideal     = round(META_AG * days_done / total_days)
re_ideal     = round(META_RE * days_done / total_days)

print(f"Taxas: qual={qual_rate:.3f}, q2a={q2a_rate:.3f}, ag2re={ag2re_rate:.3f} → META_LEADS={META_LEADS}")
print(f"Dias: done={days_done}, left={days_left}, total={total_days}")
print(f"Totais: leads={total_leads}, AG={total_ag}, RE={total_re}, NS={total_ns}, Perdidos={total_perdidos}, AGfut={total_ag_fut}")

# ── Séries para gráficos ──────────────────────────────────────────────────
wk_str   = [str(d) for d in WORKING_DAYS]
wk_label = [datetime.strptime(d,'%Y-%m-%d').strftime('%d/%m') for d in wk_str]

def build_cum(by_day, days):
    cum, s = 0, []
    for d in days:
        cum += by_day.get(str(d), 0)
        s.append(cum if d <= TODAY else None)
    return s

def build_target(meta, days):
    return [round(meta*(i+1)/len(days)) for i in range(len(days))]

# Leads acumulados (da planilha diária)
leads_series = []
cum = 0
for d in WORKING_DAYS:
    ds = str(d)
    if ds in LEADS_PLANILHA:
        cum += LEADS_PLANILHA[ds]
        leads_series.append(cum)
    elif d <= TODAY:
        leads_series.append(cum)
    else:
        leads_series.append(None)

ag_series       = build_cum(ag_by_day, WORKING_DAYS)
re_series       = build_cum(re_by_day, WORKING_DAYS)
perdidos_series = build_cum(perdidos_by_day, WORKING_DAYS)

# Perdidos por produto por dia acumulado
prods_track = ['SA', 'SF', 'PJ360']
perdidos_prod_series = {}
for p in prods_track:
    cum = 0; s = []
    for d in WORKING_DAYS:
        cum += perdidos_daily_product.get(str(d), {}).get(p, 0)
        s.append(cum if d <= TODAY else None)
    perdidos_prod_series[p] = s

leads_target = build_target(META_LEADS, WORKING_DAYS)
ag_target    = build_target(META_AG, WORKING_DAYS)
re_target    = build_target(META_RE, WORKING_DAYS)

leads_daily     = [LEADS_PLANILHA.get(str(d), None) for d in WORKING_DAYS]
perdidos_daily  = [perdidos_by_day.get(str(d), 0) if d <= TODAY else None for d in WORKING_DAYS]
loss_rate_daily = []
for l, p in zip(leads_daily, perdidos_daily):
    if l and p is not None: loss_rate_daily.append(round(p/l*100,1))
    else: loss_rate_daily.append(None)

loss_rate_avg = round(
    sum(v for v in loss_rate_daily if v is not None) /
    len([v for v in loss_rate_daily if v is not None]), 1
) if any(v is not None for v in loss_rate_daily) else 0

# ── Ordenar e limitar dicts ────────────────────────────────────────────────
def top(d, n=15): return dict(sorted(d.items(), key=lambda x:-x[1])[:n])

# ── Montar data ────────────────────────────────────────────────────────────
data = {
    'today': str(TODAY), 'updated': datetime.now().strftime('%d/%m/%Y %H:%M'),
    'days_done': days_done, 'days_left': days_left, 'total_days': total_days,
    'metas': {'leads': META_LEADS, 'ag': META_AG, 're': META_RE},
    'totais': {'leads': total_leads, 'ag': total_ag, 're': total_re,
               'ns': total_ns, 'perdidos': total_perdidos, 'ag_fut': total_ag_fut},
    'ideal_hoje': {'leads': leads_ideal, 'ag': ag_ideal, 're': re_ideal},
    'pace_proj':  {'leads': leads_proj, 'ag': ag_proj, 're': re_proj},
    'ritmo_diario': {'leads': leads_daily_needed, 'ag': ag_daily_needed, 're': re_daily_needed},
    'pace_atual': {'leads': round(leads_pace_daily), 'ag': round(ag_pace_daily,1), 're': round(re_pace_daily,1)},
    'ag_fut_ideal': ag_fut_ideal,
    'meta_closer': META_CLOSER,
    'closer_ideal_hoje': closer_ideal_hoje,
    'closer_proj':   dict(sorted(closer_proj.items(),  key=lambda x:-x[1])),
    'closer_needed': dict(sorted(closer_needed.items(), key=lambda x:-x[1])),
    'labels': wk_label,
    'series': {
        'leads': leads_series, 'leads_target': leads_target,
        'ag': ag_series, 'ag_target': ag_target,
        're': re_series, 're_target': re_target,
        'perdidos': perdidos_series,
        'leads_daily': leads_daily, 'perdidos_daily': perdidos_daily,
        'loss_rate': loss_rate_daily,
        'perdidos_SA': perdidos_prod_series.get('SA', []),
        'perdidos_SF': perdidos_prod_series.get('SF', []),
        'perdidos_PJ360': perdidos_prod_series.get('PJ360', []),
    },
    # AG realizados
    'ag_by_sdr':    top(ag_by_sdr),
    'ag_by_closer': top(ag_by_closer),
    'ag_by_product': dict(ag_by_product),
    'ag_sdr_prod':  {k: dict(v) for k,v in ag_sdr_prod.items()},
    'ag_closer_prod': {k: dict(v) for k,v in ag_closer_prod.items()},
    # RE realizadas
    're_by_closer': top(re_by_closer),
    're_by_sdr':    top(re_by_sdr),
    're_by_product': dict(re_by_product),
    # NS
    'ns_by_closer': top(ns_by_closer),
    'ns_by_sdr':    top(ns_by_sdr),
    # AG futuros
    'ag_fut_by_sdr':    top(ag_fut_by_sdr),
    'ag_fut_by_closer': top(ag_fut_by_closer),
    'ag_fut_by_product': dict(ag_fut_by_product),
    # Perdidos
    'perdidos_motivo':   dict(sorted(perdidos_by_motivo.items(), key=lambda x:-x[1])[:10]),
    'perdidos_campaign': dict(sorted(perdidos_by_campaign.items(), key=lambda x:-x[1])[:10]),
    'perdidos_etiqueta': dict(sorted(perdidos_by_etiqueta.items(), key=lambda x:-x[1])),
    # Taxas
    'loss_rate_avg': loss_rate_avg,
    'qual_rate': round(qual_rate*100,1),
    'q2a_rate':  round(q2a_rate*100,1),
    'ag2re_rate': round(ag2re_rate*100,1),
}

# ── Atualizar HTML ─────────────────────────────────────────────────────────
data_json = json.dumps(data, ensure_ascii=False, default=str)
if _os.path.exists(OUTPUT_HTML):
    with open(OUTPUT_HTML, 'r', encoding='utf-8') as f:
        html = f.read()
    html_new = re_mod.sub(r'const D = \{.*?\};', f'const D = {data_json};', html, flags=re_mod.DOTALL)
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html_new)
    print(f"\nHTML atualizado: {OUTPUT_HTML}")
else:
    json_path = _os.path.join(_os.path.dirname(OUTPUT_HTML), 'pace_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    print(f"HTML não encontrado. JSON salvo: {json_path}")
